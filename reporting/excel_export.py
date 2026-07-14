import os
import logging
from datetime import datetime
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from config.settings import REPORTS_DIR
from core.models import Trade

logger = logging.getLogger("AutoTrader")

class ExcelExporter:
    """
    Generates a professional 3-sheet Excel workbook containing:
    1. Summary Stats
    2. Detailed Trade Journal (with v6 columns: per-leg strikes, regime, phase)
    3. Daily Summary
    """
    @staticmethod
    def export_backtest(trades: List[Trade], filename: Optional[str] = None) -> str:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"backtest_report_{timestamp}.xlsx"

        filepath = REPORTS_DIR / filename
        wb = openpyxl.Workbook()

        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        
        blue_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # -------------------------------------------------------------
        # SHEET 1: Summary Metrics
        # -------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary.merge_cells("A1:C1")
        ws_summary["A1"] = "AutoTrader Backtest Performance Summary"
        ws_summary["A1"].font = title_font
        ws_summary["A1"].fill = blue_fill
        ws_summary["A1"].alignment = Alignment(horizontal="center")
        ws_summary.row_dimensions[1].height = 40

        # Calculations
        total_gross_pnl = sum(t.combined_pnl for t in trades)
        total_costs = sum(t.transaction_costs for t in trades)
        total_net_pnl = sum(t.net_pnl for t in trades)
        win_trades = [t for t in trades if t.combined_pnl > 0]
        loss_trades = [t for t in trades if t.combined_pnl <= 0]
        win_rate = (len(win_trades) / len(trades)) * 100.0 if trades else 0.0
        
        gross_profit = sum(t.combined_pnl for t in win_trades)
        gross_loss = abs(sum(t.combined_pnl for t in loss_trades))
        profit_factor = gross_profit / gross_loss if gross_loss > 0 else gross_profit

        metrics = [
            ("Metric", "Value", ""),
            ("Total Trades", len(trades), ""),
            ("Winning Trades", len(win_trades), ""),
            ("Losing Trades", len(loss_trades), ""),
            ("Win Rate", f"{win_rate:.2f}%", ""),
            ("Gross Profit", f"₹{gross_profit:,.2f}", ""),
            ("Gross Loss", f"₹{gross_loss:,.2f}", ""),
            ("Profit Factor", f"{profit_factor:.2f}", ""),
            ("Total Gross P&L", f"₹{total_gross_pnl:,.2f}", ""),
            ("Transaction Costs", f"₹{total_costs:,.2f}", ""),
            ("Net Profit/Loss", f"₹{total_net_pnl:,.2f}", "")
        ]

        for r_idx, row in enumerate(metrics, start=3):
            for c_idx, val in enumerate(row, start=1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 3:
                    cell.font = header_font
                    cell.fill = blue_fill
                    cell.alignment = Alignment(horizontal="center")
                else:
                    if c_idx == 1:
                        cell.font = bold_font
                    else:
                        cell.font = regular_font
                    
                    # Highlight Net Profit cell
                    if r_idx == 13 and c_idx == 2:
                        cell.font = bold_font
                        cell.fill = green_fill if total_net_pnl >= 0 else red_fill

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 20

        # -------------------------------------------------------------
        # SHEET 2: Trade Journal
        # -------------------------------------------------------------
        ws_journal = wb.create_sheet(title="Trade Journal")
        ws_journal.views.sheetView[0].showGridLines = True

        headers = [
            "Trade ID", "Daily SL", "Direction", "CE Strike", "PE Strike",
            "CE Entry", "PE Entry", "Entry Price (Combined)", "Qty (Lots)", "Lot Size", 
            "Regime", "Phase", "Hedge Cut Time", "Losing Leg Exit", "Losing Leg PnL", 
            "CE Exit", "PE Exit", "Exit Price (Combined)", "Exit Time", "Exit Reason", 
            "Gross PnL", "Transaction Costs", "Net PnL"
        ]

        for col_idx, h in enumerate(headers, start=1):
            cell = ws_journal.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_journal.row_dimensions[1].height = 28

        for r_idx, t in enumerate(trades, start=2):
            combined_entry = t.entry_ce_price + t.entry_pe_price
            combined_exit = (t.exit_ce_price or 0.0) + (t.exit_pe_price or 0.0) if t.exit_time else 0.0

            row_data = [
                getattr(t, "display_id", t.id),
                "POST-SL" if getattr(t, "post_daily_sl", False) else "NORMAL",
                t.direction.value,
                t.strike_ce,
                t.strike_pe,
                t.entry_ce_price,
                t.entry_pe_price,
                combined_entry,
                t.quantity,
                t.lot_size,
                t.regime_at_entry.value,
                t.phase.value,
                t.hedge_cut_time.strftime("%Y-%m-%d %H:%M:%S") if t.hedge_cut_time else "N/A",
                t.losing_leg_exit_price if t.losing_leg_exit_price is not None else "N/A",
                t.losing_leg_pnl,
                t.exit_ce_price if t.exit_ce_price is not None else "N/A",
                t.exit_pe_price if t.exit_pe_price is not None else "N/A",
                combined_exit if t.exit_time else "N/A",
                t.exit_time.strftime("%Y-%m-%d %H:%M:%S") if t.exit_time else "N/A",
                t.exit_reason.value if t.exit_reason else "OPEN",
                t.combined_pnl,
                t.transaction_costs,
                t.net_pnl
            ]

            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_journal.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_font
                
                # Format numbers
                if c_idx in [6, 7, 8, 14, 15, 16, 17, 18, 21, 22, 23]:
                    if isinstance(val, (int, float)):
                        cell.number_format = '"₹"#,##0.00'
                
                # Color Net PnL green/red
                if c_idx == 23:
                    cell.font = bold_font
                    cell.fill = green_fill if t.net_pnl >= 0 else red_fill

        # Auto-adjust column widths
        for col in ws_journal.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_journal.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # -------------------------------------------------------------
        # SHEET 3: Daily Summary
        # -------------------------------------------------------------
        ws_daily = wb.create_sheet(title="Daily Summary")
        ws_daily.views.sheetView[0].showGridLines = True

        daily_headers = ["Date", "Trades Count", "Gross PnL", "Transaction Costs", "Net PnL"]
        for col_idx, h in enumerate(daily_headers, start=1):
            cell = ws_daily.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center")

        # Group trades by date
        daily_trades = {}
        for t in trades:
            if t.entry_time:
                d_str = t.entry_time.strftime("%Y-%m-%d")
                if d_str not in daily_trades:
                    daily_trades[d_str] = []
                daily_trades[d_str].append(t)

        for r_idx, d_str in enumerate(sorted(daily_trades.keys()), start=2):
            day_trades = daily_trades[d_str]
            day_gross = sum(t.combined_pnl for t in day_trades)
            day_costs = sum(t.transaction_costs for t in day_trades)
            day_net = sum(t.net_pnl for t in day_trades)

            ws_daily.cell(row=r_idx, column=1, value=d_str).font = regular_font
            ws_daily.cell(row=r_idx, column=2, value=len(day_trades)).font = regular_font
            
            gross_cell = ws_daily.cell(row=r_idx, column=3, value=day_gross)
            gross_cell.font = regular_font
            gross_cell.number_format = '"₹"#,##0.00'

            cost_cell = ws_daily.cell(row=r_idx, column=4, value=day_costs)
            cost_cell.font = regular_font
            cost_cell.number_format = '"₹"#,##0.00'

            net_cell = ws_daily.cell(row=r_idx, column=5, value=day_net)
            net_cell.font = bold_font
            net_cell.number_format = '"₹"#,##0.00'
            net_cell.fill = green_fill if day_net >= 0 else red_fill

        ws_daily.column_dimensions["A"].width = 15
        ws_daily.column_dimensions["B"].width = 15
        ws_daily.column_dimensions["C"].width = 20
        ws_daily.column_dimensions["D"].width = 20
        ws_daily.column_dimensions["E"].width = 20

        # Save
        wb.save(filepath)
        logger.info(f"Excel report successfully generated at {filepath}")
        return str(filepath)
